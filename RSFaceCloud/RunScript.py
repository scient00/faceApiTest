#!/usr/bin/python
# -*- coding: utf-8 -*-

import logging,sys,os,shutil,yaml,string,pandas
from openpyxl import *
from RSFaceCloud.RSfaceClientCloud import RSFace
from RSFaceCloud.RSfaceOut import *
from RSFaceCloud.RSThread import *
from RSFaceCloud.RSProcess import *
from BasicMethod.BasicMethod import *
from RSFaceCloud.RSConcurrentInput import *
import datetime
logging.basicConfig(format='%(asctime)s:%(levelname)s:%(message)s', level=logging.INFO)

def PartitionSample(samefile,multiNum):
    '''
        :param samefile:输入样本列表文件
        :param multiNum:划分的样本分数
        :return:
    '''
    try:
        with open(samefile,'r') as file:
            AllData = file.readlines()
            out = []
            num = int(len(AllData)/multiNum)
            rem = int(len(AllData)%multiNum)
            for k in range(0,multiNum):
                outTemp = []
                for z in range(0,num):
                    outTemp.append(AllData[k*num + z].split('\n')[0].split('\r')[0])
                if(k==multiNum-1):
                    for x in range(0, rem):
                        outTemp.append(AllData[num * multiNum + x].split('\n')[0].split('\r')[0])
                out.append(outTemp)
            return out
    except:
        return None

def readDictionary(confpath = 'config/TestItemSetting.csv'):
    '''
    # 读取字典信息
    :param confpath:
    :return:
    '''
    try:
        pdCsv = pandas.read_csv(confpath)
        itemTest = pdCsv.set_index('SetItem').to_dict()['Value']

        return itemTest
    except:
        return {}

def MultiTest(multiNum,samefile,host,appid,appsecret,type = True):
    '''
        :param multiNum: 启动的线程数或者进程数
        :param samefile: 输入样本列表文件
        :param host: url+post
        :param appid:
        :param appsecret:
        :param type: type=True表示多进程运行,type=False表示多线程运行
        :return:并发数，总耗时，测试数据详细见dataStatistics()
    '''
    try:
        try:
            if os.path.exists(TMP_PATH):
                shutil.rmtree(TMP_PATH)
        except:
            logging.warning('delete .tmp failed!!!')
        data_statistics = vStatisticsData()
        processSample = PartitionSample(samefile, multiNum)
        multiArray = []

        startTime = datetime.datetime.now()
        for k in range(0, multiNum):
            rsInput = RsInput(k, processSample[k],host,appid,appsecret)
            if type == True:
                rsMultiArray = RsFaceProcesses(rsInput,readDictionary())
            if type == False:
                rsMultiArray = RsFaceThread(rsInput,readDictionary())
            rsMultiArray.start()
            multiArray.append(rsMultiArray)

        # 等待所有进程或线程结束
        for end in multiArray:
            end.join()
            if type == False:
                data_statistics.detect.add(end.getDataStatistics().detect)
                data_statistics.createPersonId.add(end.getDataStatistics().createPersonId)
                data_statistics.createGroups.add(end.getDataStatistics().createGroups)
                data_statistics.addFaceId.add(end.getDataStatistics().addFaceId)
                data_statistics.addPersonId.add(end.getDataStatistics().addPersonId)
                data_statistics.identification.add(end.getDataStatistics().identification)
                data_statistics.verificationByfaceId.add(end.getDataStatistics().verificationByfaceId)
                data_statistics.verificationBypersonid.add(end.getDataStatistics().verificationBypersonid)
                data_statistics.imageIdentification.add(end.getDataStatistics().imageIdentification)
                data_statistics.removeFace.add(end.getDataStatistics().removeFace)
                data_statistics.deletePersonId.add(end.getDataStatistics().deletePersonId)
                data_statistics.deleteGroups.add(end.getDataStatistics().deleteGroups)
                data_statistics.emptyPerson.add(end.getDataStatistics().emptyPerson)
                data_statistics.emptyGroups.add(end.getDataStatistics().emptyGroups)
                data_statistics.removePerson.add(end.getDataStatistics().removePerson)
        endTime = datetime.datetime.now()
        multiTotalTime = float((endTime-startTime).total_seconds()*1000)

        if type == True:
            #多进程模式读取中间数据
            with open(TMP_PATH + '/processdata.json', 'r') as rfile:
                data = json.load(rfile)
                data_statistics.push(data)

        print('测试结果\n\t: ', data_statistics)

        writeTmpData(multiNum, multiTotalTime, data_statistics, './TestResult',type)

        print("Exiting Main Thread")
        return multiNum,multiTotalTime,data_statistics
    except:
        logging.error('MultiTest')
        return 0,0.0,vStatisticsData()

def BatchMultiTest(vSampleList,host,appid,appsecret,type = True):
    '''
    :param vSampleList: 样本文件路径列表
    :param host:
    :param appid:
    :param appsecret:
    :return:
    '''
    try:
        TestResult = []
        vMultiNum = readConfFile()
        for samefile in vSampleList:
            for multiNum in vMultiNum:
                outResult = MultiTest(multiNum, samefile, host, appid, appsecret,type)
                TestResult.append(outResult)
            writeExcelReports(TestResult)
    except:
        logging.error('BatchMultiTest error!!!')

def readConfFile(confpath = 'config/IputSetting.yaml'):
    '''
    # 读取配置文件(启动的线程或者进程数量)
    :param confpath:
    :return:
    '''
    try:
        with open(confpath,'r') as cofile:
            conf = yaml.load(cofile)
        return conf.get('MULTI_THREAD_PROCESS_ARRAY')
    except:
        return []

def writeTmpData(multiNum,multiTotalTime,data_statistics = vStatisticsData(),fpath='./TestResult',type =True):
    '''
    #写入中间数据
    :param multiNum:
    :param multiTotalTime:
    :param data_statistics:
    :param fpath:
    :return:
    '''
    try:
        CreateFolder(fpath)
        outfResult = fpath + os.path.sep + '.tmpResult.ly'
        with open(outfResult, 'a') as datafile:
            if type == True:
                datafile.write("启动的进程数:\t" + str(multiNum) + '\t')
            if type == False:
                datafile.write("启动的线程数:\t" + str(multiNum) + '\n')
            datafile.write("总耗时:\t" + str(multiTotalTime) + 'ms\n')
            datafile.write('检测结果:\t'+str(data_statistics.getDict()))
            datafile.write('\n\n')
    except:
        logging.error('writeTmpData error!!!')


def writeExcelReports(TestResult,excelpath='./TestResult'):
    '''
    # 生成Excel报表
    :param excelpath:
    :return:
    '''
    try:
        dictLetter = list(string.ascii_uppercase)

        setDataList = [u'并发数',u'总耗时',u'请求数',u'请求成功数',u'请求成功总耗时',u'请求成功率',u'请求成功平均响应耗时']
        setTestItem = []
        itemTestDict = readDictionary()
        for itemIndex in itemTestDict:
            if itemTestDict[itemIndex] == 1:
                setTestItem.append(itemIndex)

        if not os.path.exists(excelpath):
            os.makedirs(excelpath)

        excelfile = excelpath + os.path.sep + str(datetime.datetime.now().date()) + '-'+str(datetime.datetime.now().hour) +'-'+ str(datetime.datetime.now().minute)+ '-'+str(datetime.datetime.now().second)  + '_FastSerachTestResult.xlsx'
        wb = Workbook()
        sheet = wb.active
        for z,title in enumerate(setTestItem):
            wb.create_sheet(index=z, title=title)

        wb.remove_sheet(wb.get_sheet_by_name('Sheet'))

        row = 2 + len(TestResult) + 2
        cols = len(['NO'] + setDataList)

        for title in setTestItem:
            sheet = wb[title]
            #sheet = wb.get_sheet_by_name(title)
            sheet.merge_cells('A1:%s1' % dictLetter[cols])
            sheet.cell(row=1, column=1, value=title)
            for k, indexName in enumerate(['NO'] + setDataList):
                sheet.cell(row=2, column=k + 1, value=indexName)

            for k, data in enumerate(TestResult):
                sheet.cell(row=k + 3, column=1, value=str(k + 1))
                sheet.cell(row=k + 3, column=2, value=data[0])
                sheet.cell(row=k + 3, column=3, value=data[1])

                sheet.cell(row=k + 3, column=4, value=data[2].getDict()[title].request_count)
                sheet.cell(row=k + 3, column=5, value=data[2].getDict()[title].request_count_success)
                sheet.cell(row=k + 3, column=6, value=data[2].getDict()[title].request_response_time)
                sheet.cell(row=k + 3, column=7, value=data[2].getDict()[title].getSucccessRate())
                sheet.cell(row=k + 3, column=8, value=data[2].getDict()[title].getAverageTime())

                outInfo = u'请求成功率:请求成功次数/请求总次数\n'
                outInfo += u'成功请求平均响应耗时:成功请求响应总耗时/请求成功次数\n'
                sheet.merge_cells('A%d:%s%d' % (row, dictLetter[cols-1], row + 4))
                sheet.cell(row=row, column=1, value=outInfo)
        wb.save(excelfile)

    except:
        logging.error('writeExcelReports error!!!')


def TestExample(host,appid,appsecret):
    '''
    #接口调用示例
    '''
    try:

        RsFace = RSFace(host, appid, appsecret)
        result = RsFace.detect('test.jpg')
        print('detect', result)
        face_id = result[0]['faces'][0]['face_id']


        result = RsFace.createPersonId(face_id, 'readsense')
        print('createPersonId', result)
        person_id = result[0]['person_id']

        #result = RsFace.deletePersonId(person_id)
        #print('deletePersonId', result)

        #person_id = '5366fa2c96e81cdb41bc05ebf43881d7'

        result = RsFace.addFaceId(face_id, person_id)
        print('addFaceId', result)


        result = RsFace.verificationByfaceId(face_id, face_id)
        print('verificationByfaceId', result)

        result = RsFace.verificationBypersonid(face_id, person_id)
        print('verificationBypersonid', result)

        # result = RsFace.removeFace(person_id,face_id)
        # print('removeFace',result)

        result = RsFace.createGroups(person_id, 'RsGroups')
        print('createGroups', result)
        groupsId = result[0]['group_id']

        result = RsFace.imageIdentification('test.jpg',groupsId)
        print('imageIdentification',result)

        result = RsFace.addPersonId(person_id, groupsId)
        print('addPersonId', result)


        result = RsFace.identification(face_id, groupsId)
        print('identification', result)

        # result = RsFace.deleteGroups(groupsId)
        # print('deleteGroups', result)

        result = RsFace.emptyPerson(person_id)
        print('emptyPerson',result)


        result = RsFace.emptyGroups(groupsId)
        print('emptyGroups', result)

        result = RsFace.removePerson(person_id, groupsId)
        print('removePerson', result)


    except Exception as ex:
        logging.exception(ex)
        logging.error('TestExample')




















